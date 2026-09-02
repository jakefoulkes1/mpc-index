"""Site-only context exporter -> data/site_context.json.

NEW, additive, site-layer only. This module writes context series the
front-end shows for ORIENTATION next to the index and the call. NONE of
these series feed the A&BG index, the market benchmark, or any locked
call - they are labelled "Context - not model inputs" on the site and
carry the same flag in the JSON.

It READS existing parsed data and Bank of England sources, and IMPORTS
(never modifies) the science-layer modules:
  - pipeline.market.ois          (polite download + OIS forward curve + SONIA)
  - pipeline.predict.market_probs (the one two-state implied-probability rule)

Three series are exported:

 (a) ois_path      - the current OIS-implied {cut, hold, hike} path for the
                     next three scheduled MPC announcements, via the SAME
                     market_probs rule used everywhere else. Meeting dates are
                     the Bank's published calendar (see UPCOMING_MEETINGS).

 (b) bank_rate_history - Bank Rate at each meeting, from data/votes.csv's
                     decided_rate column (existing decisions data), as a
                     step-chart series.

 (c) gilt_2y       - the 2-year nominal gilt yield: latest value plus a
                     12-month daily series, read from the Bank's Government
                     Liability Curve (GLC) NOMINAL files on the same Yield
                     curves page (current-month file for freshness + the
                     glcnominalddata.zip archive for history).
                     HARD STOP if the gilt files won't parse as expected -
                     never approximate (per the task instruction).

Governed by DECISIONS.md: 2026-07-12 (site v2 - context panel).

Run:  python -m pipeline.site_context
"""
import csv
import datetime as dt
import re
import time
import zipfile
from pathlib import Path

import openpyxl
import requests

from pipeline.market import ois
from pipeline.predict.market_probs import market_probs_for_meeting

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
VOTES_CSV = DATA / "votes.csv"
OUT_PATH = DATA / "site_context.json"
CACHE_DIR = DATA / "raw" / "market"

HEADERS = {"User-Agent": "mpc-index research scraper (contact: jakefoulkes@aol.com)"}
GILT_NOMINAL_ARCHIVE_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
    "yield-curves/glcnominalddata.zip"
)
# The GLC files carry the same short-end spot sheet as the OIS files.
SPOT_SHORT_SHEET = "3. spot, short end"
GILT_MATURITY_MONTHS = 24  # 2-year point
GILT_WINDOW_MONTHS = 12

# The Bank's published MPC announcement calendar, taken from
# bankofengland.co.uk/monetary-policy/upcoming-mpc-dates (that page's own
# "last updated" line read 26 May 2026 when this was transcribed). Every date
# was checked to fall on the Thursday the Bank prints beside it, and the five
# 2026 dates already in the past match the corpus's own published dates
# exactly. See DECISIONS.md, 2026-08-10.
#
# These are CONTEXT LABELS for the implied-path panel, NOT a model input and
# NOT a claim about what happens at those meetings.
#
# The Bank labels 2026 CONFIRMED and 2027 PROVISIONAL. That distinction is
# kept here rather than flattened, because a provisional date is a weaker
# claim and this file should not launder it into a firm one.
MEETINGS_2026_CONFIRMED = [
    "2026-02-05", "2026-03-19", "2026-04-30", "2026-06-18",
    "2026-07-30", "2026-09-17", "2026-11-05", "2026-12-17",
]
MEETINGS_2027_PROVISIONAL = [
    "2027-02-04", "2027-03-18", "2027-04-29", "2027-06-17",
    "2027-07-29", "2027-09-16", "2027-11-04", "2027-12-16",
]
UPCOMING_MEETINGS = MEETINGS_2026_CONFIRMED + MEETINGS_2027_PROVISIONAL

# How many still-to-come meetings the implied-path panel actually prices.
# forward_rate_for_date clips to the longest maturity the curve carries, so
# pricing every published date out to December 2027 would fill the panel with
# rows that look like data but are really the curve's endpoint repeated. Three
# is the panel's original scope and stays its scope; the calendar above is now
# the full published one so the list never runs out again.
OIS_PATH_MEETINGS = 3

DISCLAIMER = (
    "Context, not model inputs. These series are shown for orientation only; "
    "none of them feed the A&BG communication index, the market benchmark, or "
    "any locked call."
)


# ---------------------------------------------------------------------------
# (b) Bank Rate history - pure, from data/votes.csv's decided_rate column
# ---------------------------------------------------------------------------
def bank_rate_history(votes_rows: list[dict]) -> list[dict]:
    """One step point per meeting date: {date, rate_pct}. votes.csv has one
    row per member; decided_rate is the meeting's outcome, identical across
    members, so we keep one per date. decided_rate is stored as a fraction
    (0.0375 = 3.75%)."""
    by_date: dict[str, float] = {}
    for r in votes_rows:
        date = r["meeting_date"]
        by_date[date] = round(float(r["decided_rate"]) * 100, 4)
    return [{"date": d, "rate_pct": by_date[d]} for d in sorted(by_date)]


def load_votes_rows() -> list[dict]:
    with open(VOTES_CSV, newline="") as fh:
        return list(csv.DictReader(fh))


# ---------------------------------------------------------------------------
# (a) OIS-implied path for the next three meetings - reuses market_probs
# ---------------------------------------------------------------------------
def ois_path_context(curve: dict, sonia: dict,
                     meeting_dates: list[str] = None) -> dict:
    """Implied {cut, hold, hike} for each upcoming meeting, via the same
    two-state rule used for m0 and the historical benchmark."""
    meeting_dates = meeting_dates or UPCOMING_MEETINGS
    # Drop meetings that have already happened. A forward rate cannot be read
    # for a date behind the curve's own as-of date, and market_probs correctly
    # raises rather than extrapolating - which meant this whole file stopped
    # rebuilding the moment the first date in UPCOMING_MEETINGS went past
    # (2026-08-10: "meeting_date 2026-07-30 (+3d) is not after curve as_of
    # date 2026-08-07"). Filtering here keeps the panel forward-looking, which
    # is what "upcoming" already meant. It does NOT invent a replacement date:
    # the list shrinks until a real one is added from the Bank's calendar.
    # See DECISIONS.md 2026-08-10.
    as_of = dt.date.fromisoformat(curve["as_of_date"])
    upcoming = [iso for iso in meeting_dates if dt.date.fromisoformat(iso) > as_of]
    dropped = [iso for iso in meeting_dates if iso not in upcoming]
    if dropped:
        print(f"log: dropping past meeting(s) from the implied path: {dropped} "
              f"(curve as of {as_of})")
    if not upcoming:
        raise SystemExit(
            f"every date in UPCOMING_MEETINGS is on or before the curve as-of date "
            f"({as_of}). Update UPCOMING_MEETINGS in pipeline/site_context.py from "
            f"the Bank's published calendar before rebuilding."
        )
    meeting_dates = upcoming[:OIS_PATH_MEETINGS]
    meetings = []
    assumed_move_bp = lock_offset_days = None
    for iso in meeting_dates:
        p = market_probs_for_meeting(curve, sonia, dt.date.fromisoformat(iso))
        assumed_move_bp = p["assumed_move_bp"]
        lock_offset_days = p["lock_offset_days"]
        meetings.append({
            "meeting_date": p["meeting_date"],
            "forward_rate_pct": p["forward_rate_pct"],
            "implied_change_bp": p["implied_change_bp"],
            "p_cut": p["p_cut"],
            "p_hold": p["p_hold"],
            "p_hike": p["p_hike"],
        })
    return {
        "curve_as_of": curve["as_of_date"],
        "sonia_as_of": sonia["as_of_date"],
        "sonia_pct": sonia["sonia_pct"],
        "assumed_move_bp": assumed_move_bp,
        "lock_offset_days": lock_offset_days,
        "meetings": meetings,
    }


# ---------------------------------------------------------------------------
# (c) 2-year nominal gilt yield - from the Bank's GLC Nominal files
# ---------------------------------------------------------------------------
def _era_end_year(filename: str) -> int:
    """End year encoded in a GLC archive filename, e.g.
    'GLC Nominal daily data_2016 to 2024.xlsx' -> 2024,
    '..._2025 to present.xlsx' -> a large sentinel."""
    if "present" in filename.lower():
        return 9999
    years = re.findall(r"(\d{4})", filename)
    return int(years[-1]) if years else 0


def read_2y_gilt_series(xlsx_path: Path) -> list[tuple[dt.date, float]]:
    """(date, 2-year nominal spot yield %) rows from one GLC Nominal file's
    '3. spot, short end' sheet. HARD STOP if the sheet or the 24-month
    column is not present as described - never approximate."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if SPOT_SHORT_SHEET not in wb.sheetnames:
        raise ValueError(
            f"HARD STOP: '{SPOT_SHORT_SHEET}' not a sheet in {xlsx_path.name} "
            f"(sheets: {wb.sheetnames}) - gilt file structure changed."
        )
    ws = wb[SPOT_SHORT_SHEET]
    months_row = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
    col_idx = next(
        (i for i, m in enumerate(months_row)
         if isinstance(m, (int, float)) and round(m) == GILT_MATURITY_MONTHS),
        None,
    )
    if col_idx is None:
        raise ValueError(
            f"HARD STOP: no {GILT_MATURITY_MONTHS}-month maturity column in "
            f"{xlsx_path.name}::{SPOT_SHORT_SHEET} - can't locate the 2-year point."
        )
    out: list[tuple[dt.date, float]] = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        date = row[0]
        if not isinstance(date, dt.datetime):
            continue
        val = row[col_idx] if col_idx < len(row) else None
        if not isinstance(val, (int, float)):
            continue
        out.append((date.date(), round(float(val), 4)))
    wb.close()
    return out


def download_nominal_archive() -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    resp = requests.get(GILT_NOMINAL_ARCHIVE_URL, headers=HEADERS, timeout=120)
    resp.raise_for_status()
    time.sleep(2.0)
    zip_path = CACHE_DIR / "glcnominalddata.zip"
    zip_path.write_bytes(resp.content)
    extract_dir = CACHE_DIR / "glcnominalddata"
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(extract_dir)
    return extract_dir


def _gather_nominal_files() -> list[Path]:
    """Current-month GLC Nominal file (fresh, from the daily latest-yield zip
    the OIS module already downloads) + the archive era files. Only era files
    that could reach into the last 12 months are read, to stay fast."""
    latest_dir = ois.download_latest_month_zip()
    archive_dir = download_nominal_archive()
    current = sorted(latest_dir.glob("*Nominal*.xlsx"))
    if not current:
        raise ValueError(
            f"HARD STOP: no GLC Nominal current-month file in {latest_dir} "
            f"- can't get a fresh 2-year gilt yield."
        )
    # Find the freshest date from the current-month file to bound which
    # archive eras are worth reading.
    latest_rows = read_2y_gilt_series(current[0])
    if not latest_rows:
        raise ValueError(f"HARD STOP: {current[0].name} yielded no 2-year gilt rows.")
    latest_year = max(d for d, _ in latest_rows).year
    archive = [
        p for p in sorted(archive_dir.glob("*Nominal*.xlsx"))
        if _era_end_year(p.name) >= latest_year - 1
    ]
    return archive + current


def gilt_2y_context(series: list[tuple[dt.date, float]] = None) -> dict:
    """Latest 2-year nominal gilt yield + a 12-month daily series. Pass
    `series` (merged, sorted (date, yield)) to test without a live call;
    otherwise gathers and reads the Bank's GLC Nominal files."""
    if series is None:
        merged: dict[dt.date, float] = {}
        for path in _gather_nominal_files():
            for date, val in read_2y_gilt_series(path):
                merged[date] = val  # later files (fresher) win on overlap
        series = sorted(merged.items())
    if not series:
        raise ValueError("HARD STOP: no 2-year gilt data parsed from any GLC Nominal file.")

    latest_date, latest_val = series[-1]
    cutoff = latest_date - dt.timedelta(days=365)
    window = [(d, v) for d, v in series if d >= cutoff]
    return {
        "label": "2-year nominal gilt yield",
        "source": "Bank of England GLC Nominal curve, '3. spot, short end', 24-month point",
        "as_of": latest_date.isoformat(),
        "latest_pct": latest_val,
        "window_months": GILT_WINDOW_MONTHS,
        "sparkline": [{"date": d.isoformat(), "yield_pct": v} for d, v in window],
    }


# ---------------------------------------------------------------------------
# assembly
# ---------------------------------------------------------------------------
def build_context(curve: dict, sonia: dict, votes_rows: list[dict],
                  gilt: dict, generated_utc: str = None) -> dict:
    generated_utc = generated_utc or dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
    return {
        "schema": "site-context-v1",
        "generated_utc": generated_utc,
        "context_not_model_inputs": True,
        "disclaimer": DISCLAIMER,
        "ois_path": ois_path_context(curve, sonia),
        "bank_rate_history": {
            "source": "data/votes.csv decided_rate (existing decisions data)",
            "points": bank_rate_history(votes_rows),
        },
        "gilt_2y": gilt,
        "sources": {
            "ois_curve": ois.OIS_ZIP_URL,
            "ois_latest_month": ois.LATEST_MONTH_ZIP_URL,
            "sonia_series_code": ois.SONIA_SERIES_CODE,
            "gilt_nominal_archive": GILT_NOMINAL_ARCHIVE_URL,
        },
    }


def main() -> None:
    import json
    print("Fetching OIS forward curve + SONIA (live) ...")
    curve = ois.latest_forward_curve()
    sonia = ois.latest_sonia()
    print(f"  curve as of {curve['as_of_date']}, SONIA {sonia['sonia_pct']:.4f}% "
          f"as of {sonia['as_of_date']}")
    print("Reading 2-year nominal gilt yield (current month + archive) ...")
    gilt = gilt_2y_context()
    print(f"  2-year gilt {gilt['latest_pct']:.4f}% as of {gilt['as_of']} "
          f"({len(gilt['sparkline'])} days in the 12-month sparkline)")
    votes_rows = load_votes_rows()
    context = build_context(curve, sonia, votes_rows, gilt)
    OUT_PATH.write_text(json.dumps(context, indent=2) + "\n")
    print(f"Wrote {OUT_PATH.relative_to(ROOT)}")
    for m in context["ois_path"]["meetings"]:
        print(f"  {m['meeting_date']}: implied {m['implied_change_bp']:+.2f}bp -> "
              f"cut {m['p_cut']:.2f} / hold {m['p_hold']:.2f} / hike {m['p_hike']:.2f}")


if __name__ == "__main__":
    main()
