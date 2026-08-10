"""Parse the Bank's mpcvoting.xlsx into data/votes.csv.

One row per (meeting, member): the member's preferred Bank Rate against the
decided rate, plus meeting-level skew and dissent counts (repeated on every
row for that meeting - a standard denormalised CSV, easy to load and group).

skew = average(preferred rates of all voting members) - decided rate,
following Apel & Blix Grimaldi (2012) p.13 (after Gerlach-Kristen 2004):
skew = average(r_j) - r. Positive skew = committee leaned for a higher rate
than decided (hawkish dissent); negative = leaned lower (dovish dissent).

The sheet's own date column is the meeting's PUBLISHED/announcement date,
not the meeting_end date - confirmed by cross-checking known dates (e.g.
2026-06-18 matches minutes-2026-06's `published` field, not its
`meeting_end` of 2026-06-17). Rates are decimals (0.0375 = 3.75%), kept as
given in the source.

Governed by DECISIONS.md: 2026-07-11 (voting-record parsing and
reconciliation).

Run:  python -m pipeline.build_votes
"""
import csv
import datetime as dt
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "data" / "raw" / "mpcvoting.xlsx"
INDEX_PATH = ROOT / "data" / "index.json"
OUT = ROOT / "data" / "votes.csv"

HEADER_ROW = 4
LABEL_COLUMNS = {"Current members", "Past members"}


def corpus_window() -> tuple[dt.date, dt.date]:
    """The era the voting sheet is filtered to, derived from the text corpus.

    Returns (start, end_exclusive) spanning exactly the corpus's published
    dates. This used to be two hardcoded constants that had to be bumped by
    hand every ingest; on 2026-08-10 the July meeting was silently dropped
    from votes.csv because the bound still said 2026-07-01, and the script
    reported 94 meetings with no error. Deriving it makes that failure mode
    impossible: the voting sheet is filtered to the corpus, so a meeting in
    the corpus can never fall outside the window.

    Also still sidesteps the pre-2015 rows that record a dissent as
    qualitative "Increase"/"Decrease" text with no rate (e.g. 1998), since
    the corpus starts in August 2015.

    See DECISIONS.md 2026-08-10.
    """
    if not INDEX_PATH.exists():
        raise SystemExit(
            f"{INDEX_PATH} not found - the voting window is derived from the corpus, "
            f"so build the index first: python -m pipeline.build_index"
        )
    corpus = json.loads(INDEX_PATH.read_text())
    published = sorted(d["published"] for d in corpus["documents"] if d["published"])
    if not published:
        raise SystemExit(f"{INDEX_PATH} has no published documents - cannot derive a window")
    start = dt.date.fromisoformat(published[0])
    end = dt.date.fromisoformat(published[-1]) + dt.timedelta(days=1)
    return start, end


def load_member_columns(ws) -> dict[int, str]:
    columns = {}
    for cell in ws[HEADER_ROW]:
        if cell.value and cell.value not in LABEL_COLUMNS:
            name = " ".join(str(cell.value).split())  # collapse embedded newlines
            columns[cell.column - 1] = name
    return columns


def parse_meetings(ws, member_columns: dict[int, str],
                   era_start: dt.date, era_end: dt.date) -> list[dict]:
    meetings = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        date, decided = row[1], row[2]
        if not isinstance(date, dt.datetime) or not isinstance(decided, (int, float)):
            continue
        if not (era_start <= date.date() < era_end):
            continue
        votes = {}
        for i, name in member_columns.items():
            v = row[i]
            if v is None:
                continue
            if not isinstance(v, (int, float)):
                print(f"log: {date.date()}: {name}'s vote is non-numeric ({v!r}) - excluded from skew/dissent counts")
                continue
            votes[name] = v
        if not votes:
            continue
        preferred = list(votes.values())
        skew = sum(preferred) / len(preferred) - decided
        hawkish_dissents = sum(1 for r in preferred if r > decided)
        dovish_dissents = sum(1 for r in preferred if r < decided)
        meetings.append({
            "meeting_date": date.date().isoformat(),
            "decided_rate": decided,
            "skew": round(skew, 6),
            "hawkish_dissents": hawkish_dissents,
            "dovish_dissents": dovish_dissents,
            "votes": votes,
        })
    return meetings


def reconcile_against_corpus(meetings: list[dict], era_start: dt.date, era_end: dt.date) -> None:
    if not INDEX_PATH.exists():
        print("no index.json to reconcile against - skipping reconciliation")
        return
    corpus = json.loads(INDEX_PATH.read_text())
    corpus_published = {d["published"] for d in corpus["documents"] if d["published"]}
    sheet_dates = {m["meeting_date"] for m in meetings}

    start, end = era_start.isoformat(), era_end.isoformat()
    corpus_only = sorted(d for d in corpus_published if start <= d < end and d not in sheet_dates)
    sheet_only = sorted(d for d in sheet_dates if start <= d < end and d not in corpus_published)

    print(f"reconciliation (published date == voting-sheet meeting date, {start} to {end}):")
    print(f"  in corpus, no matching voting-sheet row ({len(corpus_only)}): {corpus_only}")
    print(f"  in voting sheet, no matching corpus document ({len(sheet_only)}): {sheet_only}")


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Bank Rate Decisions"]
    member_columns = load_member_columns(ws)
    era_start, era_end = corpus_window()
    meetings = parse_meetings(ws, member_columns, era_start, era_end)

    with open(OUT, "w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["meeting_date", "decided_rate", "member", "preferred_rate",
                          "skew", "hawkish_dissents", "dovish_dissents"])
        for m in meetings:
            for member, preferred in sorted(m["votes"].items()):
                writer.writerow([m["meeting_date"], m["decided_rate"], member, preferred,
                                  m["skew"], m["hawkish_dissents"], m["dovish_dissents"]])

    print(f"wrote {OUT} ({len(meetings)} meetings, {sum(len(m['votes']) for m in meetings)} member-votes)")
    reconcile_against_corpus(meetings, era_start, era_end)


if __name__ == "__main__":
    main()
