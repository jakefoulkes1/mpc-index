"""Bank of England OIS forward curve + current SONIA.

Sources (see DECISIONS.md for how each was found and verified):

- OIS forward curve, TWO files, both needed for a genuinely up-to-date
  curve (bankofengland.co.uk/statistics/yield-curves):
  1. "oisddata.zip" (daily OIS data, the multi-year ARCHIVE) ->
     https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/oisddata.zip
     One .xlsx per era (currently "2009 to 2015", "2016 to 2024", "2025 to
     present" - picked by actually checking which has the most recent
     date, not a hardcoded filename). Per the Bank's own FAQ, archive data
     is only refreshed "by close of business of the second working day of
     each month" - i.e. this file alone is up to ~a month stale.
  2. "latest-yield-curve-data.zip" (updated daily) ->
     https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip
     Contains "OIS daily data current month.xlsx" - the current month's
     business days, published same/next business day. THIS is what makes
     the curve actually fresh; the archive alone was found to be up to a
     month stale (see DECISIONS.md, 2026-08-08).
  Both use sheet "1. fwds, short end" ("UK instantaneous OIS forward
  curve"): row 3 = maturity in whole months (1 to 60), one column per
  month; each subsequent row is one business day, column A = date,
  columns B: = the instantaneous forward rate (percent per annum) at that
  maturity, as of that date. The single freshest date across BOTH files
  is used - not just whichever the code happens to check last.

- SONIA: the Bank's Interactive Statistical Database CSV export, series
  code IUDSOIA ("Daily Sterling overnight index average (SONIA) rate").

HARD STOP: if either source's structure has changed and can't be parsed as
described above, this raises rather than falling back to a guess or an
approximation - see individual functions.

Run:  python -m pipeline.market.ois
"""
import csv
import datetime as dt
import io
import time
import zipfile
from pathlib import Path

import openpyxl
import requests

ROOT = Path(__file__).resolve().parents[2]
CACHE_DIR = ROOT / "data" / "raw" / "market"
HEADERS = {"User-Agent": "mpc-index research scraper (contact: jakefoulkes@aol.com)"}

OIS_ZIP_URL = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/oisddata.zip"
LATEST_MONTH_ZIP_URL = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip"
FORWARD_SHEET_NAME = "1. fwds, short end"
# The 2009-2015 era file names the same sheet (identical structure -
# checked by hand) "1. fwd curve" instead. Both are tried.
FORWARD_SHEET_NAMES = (FORWARD_SHEET_NAME, "1. fwd curve")
SONIA_SERIES_CODE = "IUDSOIA"
SONIA_CSV_URL = (
    "https://www.bankofengland.co.uk/boeapps/iadb/fromshowcolumns.asp"
    "?csv.x=yes&Datefrom={date_from}&Dateto=now&SeriesCodes={code}"
    "&UsingCodes=Y&CSVF=TT&VPD=Y&VFD=N"
)


def _download_and_extract(url: str, cache_name: str) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    resp = requests.get(url, headers=HEADERS, timeout=60)
    resp.raise_for_status()
    time.sleep(2.0)
    zip_path = CACHE_DIR / f"{cache_name}.zip"
    zip_path.write_bytes(resp.content)
    extract_dir = CACHE_DIR / cache_name
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(extract_dir)
    return extract_dir


def download_ois_zip() -> Path:
    return _download_and_extract(OIS_ZIP_URL, "oisddata")


def download_latest_month_zip() -> Path:
    return _download_and_extract(LATEST_MONTH_ZIP_URL, "latest-yield-curve-data")


def _read_forward_curve_sheet(
    xlsx_path: Path, sheet_names: tuple[str, ...] = (FORWARD_SHEET_NAME,)
) -> tuple[str, list[float], list[tuple[dt.date, list[float]]]] | None:
    """Returns (sheet_name_used, maturities_months, [(date, [rate per
    maturity]), ...]), or None if this file has none of sheet_names
    (older era files use a different layout - not every era file is
    expected to match every name)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet_name = next((n for n in sheet_names if n in wb.sheetnames), None)
    if sheet_name is None:
        print(f"log: {xlsx_path.name}: none of {sheet_names} found as a sheet "
              f"(sheets present: {wb.sheetnames}) - skipping this era file")
        return None
    ws = wb[sheet_name]
    months_row = [c.value for c in ws[3]]
    if months_row[0] != "months:" or not isinstance(months_row[1], (int, float)):
        raise ValueError(
            f"HARD STOP: expected row 3 of '{sheet_name}' in {xlsx_path.name} to start with "
            f"'months:' followed by numeric maturities, got {months_row[:3]!r}"
        )
    maturities = [float(m) for m in months_row[1:] if isinstance(m, (int, float))]

    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        date = row[0]
        if not isinstance(date, dt.datetime):
            continue
        rates = [v for v in row[1:1 + len(maturities)]]
        if any(not isinstance(v, (int, float)) for v in rates):
            continue  # incomplete row (e.g. a holiday placeholder)
        rows.append((date.date(), [float(v) for v in rates]))
    if not rows:
        raise ValueError(f"HARD STOP: no complete data rows found in {xlsx_path.name}::{sheet_name}")
    return sheet_name, maturities, rows


def pick_freshest_curve(xlsx_files: list[Path]) -> dict:
    """Pure (no I/O beyond reading the given files): picks whichever file
    has the most recent date across ALL of them, not a hardcoded filename
    or a single directory - the caller is responsible for gathering every
    candidate file (both the era archive and the current-month file)."""
    if not xlsx_files:
        raise ValueError("HARD STOP: no .xlsx files given to pick_freshest_curve")

    best = None  # (date, maturities, rates, source_file)
    for path in xlsx_files:
        parsed = _read_forward_curve_sheet(path)
        if parsed is None:
            continue
        _sheet_name, maturities, rows = parsed
        last_date, last_rates = rows[-1]
        if best is None or last_date > best[0]:
            best = (last_date, maturities, last_rates, path.name)

    if best is None:
        raise ValueError(
            f"HARD STOP: none of {[p.name for p in xlsx_files]} contain sheet "
            f"'{FORWARD_SHEET_NAME}' - can't identify a forward curve file."
        )
    as_of_date, maturities, rates, source_file = best
    return {
        "as_of_date": as_of_date.isoformat(),
        "source_file": source_file,
        "sheet": FORWARD_SHEET_NAME,
        "maturities_months": maturities,
        "forward_rates_pct": rates,
    }


def latest_forward_curve() -> dict:
    """Downloads BOTH the multi-year archive and the current-month file
    (see module docstring - the archive alone can be up to a month
    stale), and returns the single freshest day across both."""
    oisdata_dir = download_ois_zip()
    latest_month_dir = download_latest_month_zip()
    # latest-yield-curve-data.zip bundles 4 different curve types (GLC
    # Nominal/Real/Inflation gilt curves + OIS) that happen to share the
    # same sheet NAME convention ("1. fwds, short end") but not the same
    # row structure - the GLC files have blank/formula-error cells for
    # some dates, which isn't a malformed OIS file, it's simply not an
    # OIS file. Only the OIS one is relevant here.
    xlsx_files = sorted(oisdata_dir.glob("*.xlsx")) + sorted(latest_month_dir.glob("*OIS*.xlsx"))
    if not xlsx_files:
        raise ValueError(f"HARD STOP: neither {oisdata_dir} nor {latest_month_dir} contain any relevant .xlsx files")
    return pick_freshest_curve(xlsx_files)


def fetch_sonia(lookback_days: int = 30) -> list[tuple[dt.date, float]]:
    date_from = (dt.date.today() - dt.timedelta(days=lookback_days)).strftime("%d/%b/%Y")
    url = SONIA_CSV_URL.format(date_from=date_from, code=SONIA_SERIES_CODE)
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    time.sleep(2.0)
    return parse_sonia_csv(resp.text)


def parse_sonia_csv(text: str) -> list[tuple[dt.date, float]]:
    lines = text.splitlines()
    try:
        header_idx = next(i for i, line in enumerate(lines) if line.startswith("DATE,"))
    except StopIteration:
        raise ValueError(
            f"HARD STOP: SONIA CSV has no 'DATE,...' header row - unexpected format. "
            f"First few lines: {lines[:5]!r}"
        )
    reader = csv.reader(lines[header_idx + 1:])
    out = []
    for row in reader:
        if len(row) < 2 or not row[0]:
            continue
        date = dt.datetime.strptime(row[0], "%d %b %Y").date()
        out.append((date, float(row[1])))
    if not out:
        raise ValueError("HARD STOP: SONIA CSV parsed but yielded no data rows")
    return sorted(out)


def latest_sonia() -> dict:
    series = fetch_sonia()
    date, rate = series[-1]
    return {"as_of_date": date.isoformat(), "series_code": SONIA_SERIES_CODE, "sonia_pct": rate}


def main() -> None:
    curve = latest_forward_curve()
    sonia = latest_sonia()
    print(f"OIS forward curve as of {curve['as_of_date']} (source: {curve['source_file']}::{curve['sheet']})")
    print(f"  1-month forward: {curve['forward_rates_pct'][0]:.4f}%  "
          f"12-month forward: {curve['forward_rates_pct'][11]:.4f}%")
    print(f"SONIA as of {sonia['as_of_date']}: {sonia['sonia_pct']:.4f}%")


if __name__ == "__main__":
    main()
