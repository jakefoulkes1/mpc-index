import datetime as dt

import openpyxl
import pytest

from pipeline.market.ois import _read_forward_curve_sheet, parse_sonia_csv, pick_freshest_curve

SONIA_FIXTURE = """SERIES,DESCRIPTION
IUDSOIA,Daily Sterling overnight index average (SONIA) rate

DATE,IUDSOIA
01 Jun 2026,3.7291
02 Jun 2026,3.7308
03 Jun 2026,3.7306
"""


def _sample_forward_curve_workbook(tmp_path, sheet_name="1. fwds, short end"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([None, "UK instantaneous OIS forward curve"])
    ws.append(["Maturity "])
    ws.append(["months:", 1, 2, 3])
    ws.append(["years:", 1 / 12, 2 / 12, 3 / 12])
    ws.append(["Refresh", "Refresh", "Refresh", "Refresh"])
    ws.append([dt.datetime(2026, 6, 29), 3.72, 3.76, 3.80])
    ws.append([dt.datetime(2026, 6, 30), 3.73, 3.77, 3.81])
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


def test_read_forward_curve_sheet_happy_path(tmp_path):
    path = _sample_forward_curve_workbook(tmp_path)
    sheet_name, maturities, rows = _read_forward_curve_sheet(path)
    assert sheet_name == "1. fwds, short end"
    assert maturities == [1.0, 2.0, 3.0]
    assert rows[-1] == (dt.date(2026, 6, 30), [3.73, 3.77, 3.81])


def test_read_forward_curve_sheet_tries_alternate_sheet_names(tmp_path):
    path = _sample_forward_curve_workbook(tmp_path, sheet_name="1. fwd curve")
    sheet_name, maturities, rows = _read_forward_curve_sheet(
        path, sheet_names=("1. fwds, short end", "1. fwd curve")
    )
    assert sheet_name == "1. fwd curve"
    assert maturities == [1.0, 2.0, 3.0]


def test_read_forward_curve_sheet_returns_none_for_wrong_layout(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "some other sheet"
    path = tmp_path / "wrong.xlsx"
    wb.save(path)
    assert _read_forward_curve_sheet(path) is None


def _workbook_with_last_date(tmp_path, filename, last_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "1. fwds, short end"
    ws.append([None, "UK instantaneous OIS forward curve"])
    ws.append(["Maturity "])
    ws.append(["months:", 1, 2])
    ws.append(["years:", 1 / 12, 2 / 12])
    ws.append(["Refresh", "Refresh", "Refresh"])
    ws.append([last_date - dt.timedelta(days=1), 3.70, 3.74])
    ws.append([last_date, 3.71, 3.75])
    path = tmp_path / filename
    wb.save(path)
    return path


def test_pick_freshest_curve_prefers_the_more_recent_file(tmp_path):
    # Simulates the real bug: an "archive" file that's stale next to a
    # "current month" file with newer data - the freshest date across
    # BOTH must win, not just whichever file is checked first.
    stale = _workbook_with_last_date(tmp_path, "archive.xlsx", dt.date(2026, 6, 30))
    fresh = _workbook_with_last_date(tmp_path, "current_month.xlsx", dt.date(2026, 7, 9))
    result = pick_freshest_curve([stale, fresh])
    assert result["as_of_date"] == "2026-07-09"
    assert result["source_file"] == "current_month.xlsx"


def test_pick_freshest_curve_hard_stops_on_empty_list():
    with pytest.raises(ValueError, match="HARD STOP"):
        pick_freshest_curve([])


def test_parse_sonia_csv():
    rows = parse_sonia_csv(SONIA_FIXTURE)
    assert rows == [
        (dt.date(2026, 6, 1), 3.7291),
        (dt.date(2026, 6, 2), 3.7308),
        (dt.date(2026, 6, 3), 3.7306),
    ]


def test_parse_sonia_csv_hard_stops_on_missing_header():
    with pytest.raises(ValueError, match="HARD STOP"):
        parse_sonia_csv("not,a,sonia,csv\n1,2,3")
